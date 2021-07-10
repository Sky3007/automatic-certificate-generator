import openpyxl  #to read file from excel it is not gaurd against quadratic 
                 #blow and billions use defusegaurd to gaurd.
from PIL import Image  #python image library
from PIL import ImageFont
from PIL import ImageDraw

import numpy

import cv2 as cv

import os

import smtplib

import glob

from email import encoders

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase

l=0
def certificate():
    coordinate_y_adjustment = 84

    coordinate_x_adjustment = 5
    path="C:\\Users\\anjal\\OneDrive\\Desktop\\certificate\\namelist.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active 
    #cell_o=sheet_obj.cell(row=2,column=2)
    #print(cell_o)
    font_size = 2
    font_color = (0,0,255,255)#blue
    dm=(680,580)
    img=Image.open(r"C:\Users\anjal\OneDrive\Desktop\certificate\ck.jpg")
    p=Image.open(r"C:\Users\anjal\OneDrive\Desktop\certificate\sig.jpg")

    back=img.copy()
    back.paste(p, (800, 600))
    back.save('copyP.jpg', quality=95)
    ig=Image.open("copyP.jpg")

    fnt = ImageFont.truetype('arial.ttf', 50)
    d = ImageDraw.Draw(ig)
    d.text((165,645), "10/02/19", font=fnt, fill=(0, 0, 0))
    ig.save('pk.jpg')
    
    i=0
    images=[]
    for x in range (1,9):
        get_name = sheet_obj.cell(row = x ,column = 1) 
        certi_name = get_name.value 
        get = sheet_obj.cell(row = x ,column = 2) 
        certi = get.value 
        img2 = cv.imread('pk.jpg',0)
        img=cv.resize(img2,dm)
        font =cv.FONT_HERSHEY_SCRIPT_COMPLEX
        text_size = cv.getTextSize(certi_name, font, font_size, 2)[0] 

        x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_adjustment  
        y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_adjustment 
        x = int(x) 
        y = int(y)   
        cv.putText(img, certi_name, 
              (x ,y ),  
              font, 
              font_size, 
              font_color, 3) 
    
        path = 'C://Users//anjal//OneDrive//Desktop//certificate//imgg'
    
        cv.imshow("image",img)
        cv.imwrite("C://Users//anjal//OneDrive//Desktop//certificate//imgg//image."+str(i)+".jpg",img)
        i=i+1
        print(i)
        k=cv.waitKey(0)

def gmail():
    l=0
    path="C:\\Users\\anjal\\OneDrive\\Desktop\\certificate\\namelist.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    for x in range(1,5):
        get = sheet_obj.cell(row = x ,column = 2) 
        certi = get.value 
        mail_content = '''Hello,
    This is a test mail.
    In this mail we are sending your Achivment certifiate.
    You did a great Job.
    Thank You
    '''     
    sender_address = 'anjali.goyal3007@gmail.com'
    sender_pass = 'maalove@123@'
    #receiver_address = 'anjaligoyal730@gmail.com'
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = certi
    message['Subject'] = 'A test mail sent by BOSS. It has an attachment.'

    message.attach(MIMEText(mail_content, 'plain'))
    attach_file_name = "C://Users//anjal//OneDrive//Desktop//certificate//imgg//image.%x.jpg"%l
    attach_file = open(attach_file_name, 'rb') # Open the file as binary mode
    payload = MIMEBase('application', 'octate-stream')
    payload.set_payload((attach_file).read())
    encoders.encode_base64(payload) #encode the attachment
    #add payload header with filename
    payload.add_header('Content-Decomposition', 'attachment', filename=attach_file_name)
    message.attach(payload)
    #Create SMTP session for sending the mail
    session = smtplib.SMTP('smtp.gmail.com', 587) #use gmail with port
    session.starttls() #enable security
    session.login(sender_address, sender_pass) #login with mail_id and password
    text = message.as_string()
    session.sendmail(sender_address,certi, text)
    session.quit()
    print('Mail Sent')
    l+=1
print("DONE")

certificate()
gmail()
